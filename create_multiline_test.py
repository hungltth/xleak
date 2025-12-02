#!/usr/bin/env python3
"""Create a test Excel file with multi-line cell content."""
import openpyxl

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MultilineTest"

# Add header
ws['A1'] = "Number"
ws['B1'] = "Description"

# Add cell with many lines (like the issue describes)
multiline_numbers = "\n".join(str(i) for i in range(1, 21))
ws['A2'] = multiline_numbers

# Add another multi-line cell
ws['B2'] = "This is a cell\nwith multiple lines\nof text content\nto test\nthe scrolling\nfunctionality\nin the cell\ndetail popup\nview.\nLine 10\nLine 11\nLine 12\nLine 13\nLine 14\nLine 15"

# Save
wb.save('tests/fixtures/multiline_test.xlsx')
print("Created tests/fixtures/multiline_test.xlsx with multi-line cells")
